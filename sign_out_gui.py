
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook

# Load the Excel file
file_path = 'Copy of library books log 10TH MARCH 2.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# Function to sign out a book
def sign_out_book(isbn):
    # Find the row with the given ISBN
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == isbn:
            # Get the current date
            borrow_date = datetime.now().date()
            # Calculate the return date (14 days later)
            return_date = borrow_date + timedelta(days=14)
            # Update the Excel file
            ws.cell(row=row[0], column=6, value=borrow_date.strftime('%Y-%m-%d'))
            ws.cell(row=row[0], column=7, value=return_date.strftime('%Y-%m-%d'))
            wb.save(file_path)
            return borrow_date, return_date
    return None, None

# Function to handle the sign out process
def handle_sign_out():
    isbn = isbn_entry.get()
    if isbn:
        borrow_date, return_date = sign_out_book(isbn)
        if borrow_date and return_date:
            messagebox.showinfo("Success", f"Book signed out!\nBorrow Date: {borrow_date}\nReturn Date: {return_date}")
        else:
            messagebox.showerror("Error", "ISBN not found!")
    else:
        messagebox.showerror("Error", "Please enter an ISBN!")

# Create the GUI
root = tk.Tk()
root.title("Library Book Sign Out")

tk.Label(root, text="Enter ISBN:").grid(row=0, column=0, padx=10, pady=10)
isbn_entry = tk.Entry(root)
isbn_entry.grid(row=0, column=1, padx=10, pady=10)

sign_out_button = tk.Button(root, text="Sign Out Book", command=handle_sign_out)
sign_out_button.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()
